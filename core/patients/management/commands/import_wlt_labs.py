"""
Management command: import_wlt_labs
=====================================
Imports ALT/AST/Bilirubin lab timepoints for WLT patients from
WLT_PATIENTS_cleaned.xlsx.

import_patients.py reads WLT_PATIENTS_cleaned.txt, a fixed-width text export
that lost the lab columns during conversion (demographics/scores only). The
.xlsx sibling of that same export still has them, structured identically to
the SLT source (ALT/AST/TB at preop/week1/month1/year1) — this command
matches existing WLT patients by patient_id and fills in those timepoints.

Usage (from core/ directory, with venv active):
    python manage.py import_wlt_labs path/to/WLT_PATIENTS_cleaned.xlsx
"""
import os
from django.core.management.base import BaseCommand, CommandError
from patients.models import Patient, PatientTimepoint

# Column indices (0-based) in the "Patients" sheet
COL_PATIENT_ID = 0
COL_LABS = {
    "preop":  (23, 24, 25),
    "week1":  (26, 27, 28),
    "month1": (29, 30, 31),
    "year1":  (32, 33, 34),
}


def _num(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


class Command(BaseCommand):
    help = "Import ALT/AST/Bilirubin lab timepoints for WLT patients from WLT_PATIENTS_cleaned.xlsx"

    def add_arguments(self, parser):
        parser.add_argument("filepath", help="Path to WLT_PATIENTS_cleaned.xlsx")

    def handle(self, *args, **options):
        fp = options["filepath"]
        if not os.path.exists(fp):
            raise CommandError(f"File not found: {fp}")

        import openpyxl
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        ws = wb["Patients"] if "Patients" in wb.sheetnames else wb.worksheets[0]

        matched = skipped_no_patient = timepoints_upserted = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[COL_PATIENT_ID]:
                continue
            pid = str(row[COL_PATIENT_ID]).strip()

            try:
                patient = Patient.objects.get(patient_id=pid)
            except Patient.DoesNotExist:
                skipped_no_patient += 1
                continue
            matched += 1

            for phase, (alt_i, ast_i, tb_i) in COL_LABS.items():
                alt = _num(row[alt_i])
                ast = _num(row[ast_i])
                tb  = _num(row[tb_i])
                if alt is None and ast is None and tb is None:
                    continue
                PatientTimepoint.objects.update_or_create(
                    patient=patient, timepoint=phase,
                    defaults={"alt": alt, "ast": ast, "bilirubin": tb},
                )
                timepoints_upserted += 1

        self.stdout.write(self.style.SUCCESS(
            f"\nDone. Patients matched: {matched}  Skipped (no matching patient): {skipped_no_patient}"
            f"\nTimepoints upserted: {timepoints_upserted}"
        ))
